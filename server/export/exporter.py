"""Filtered CSV / XLSX export of both patient datasets.

Two stores feed this: intake_sessions (the reception form) and
consultation_notes (the scribe's summaries). Both keep their real content in
nested JSON, which is right for the app but useless in a spreadsheet — so each
dataset gets a flattener that pulls the JSON out into ordered, human-labelled
columns. Everything downstream (filtering, CSV, XLSX) is dataset-agnostic and
works off those flat rows.

Read-only: this module only ever SELECTs. It never writes, so exporting the
chatbot's intake data can't affect the chatbot.
"""
import csv
import io
from dataclasses import dataclass, field
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from services.db import get_client

# Excel's hard per-cell limit. A long consultation transcript is the only field
# that can approach it; truncate rather than produce a corrupt workbook.
_CELL_MAX = 32000

DATASETS = ("intake", "consultations")
_TABLE = {"intake": "intake_sessions", "consultations": "consultation_notes"}


@dataclass
class ExportFilters:
    date_from: str | None = None  # ISO date/datetime; inclusive
    date_to: str | None = None  # ISO date/datetime; inclusive (whole day if bare date)
    statuses: list[str] = field(default_factory=list)
    ids: list[str] = field(default_factory=list)  # specific rows ("selected patients")
    flagged_only: bool = False  # intake only — sessions that raised a flag
    search: str = ""  # case-insensitive substring over name/phone/complaint


# --- flattening --------------------------------------------------------------


def _get(node, dotted: str):
    for part in dotted.split("."):
        if not isinstance(node, dict):
            return None
        node = node.get(part)
    return node


def _join(value, sep: str = "; ") -> str:
    """Render a list/dict/scalar from the JSON into one readable cell string."""
    if value is None:
        return ""
    if isinstance(value, list):
        parts = []
        for item in value:
            if isinstance(item, dict):
                parts.append(" ".join(str(v) for v in item.values() if v not in (None, "")))
            elif item not in (None, ""):
                parts.append(str(item))
        return sep.join(parts)
    if isinstance(value, dict):
        return sep.join(f"{k}: {v}" for k, v in value.items() if v not in (None, ""))
    return str(value)


def _flatten_intake(row: dict) -> dict:
    p = row.get("patient") or {}
    flags = row.get("flags") or []
    return {
        "ID": row.get("id"),
        "Created": row.get("created_at"),
        "Status": row.get("status"),
        "Full name": _get(p, "personal.full_name"),
        "Age": _get(p, "personal.age"),
        "Gender": _get(p, "personal.gender"),
        "Phone": _get(p, "personal.phone"),
        "Address": _get(p, "personal.address"),
        "Symptoms": _join(_get(p, "visit.symptoms")),
        "Allergies": _join(_get(p, "medical_history.allergies")),
        "Existing conditions": _join(_get(p, "medical_history.existing_conditions")),
        "Current medications": _join(_get(p, "medical_history.current_medications")),
        "Flags": _join([f"{f.get('field')}: {f.get('reason')}" for f in flags]),
        "Language": _get(p, "meta.language"),
    }


def _flatten_consultation(row: dict) -> dict:
    s = row.get("summary") or {}
    meds = s.get("medications_prescribed") or []
    med_str = "; ".join(
        " · ".join(str(x) for x in (m.get("name"), m.get("dosage"), m.get("frequency"), m.get("duration")) if x)
        for m in meds
        if isinstance(m, dict)
    )
    transcript = " | ".join(
        str(seg.get("text", "")) for seg in (row.get("transcript") or []) if isinstance(seg, dict)
    )
    return {
        "ID": row.get("id"),
        "Created": row.get("created_at"),
        "Status": row.get("status"),
        "Duration (s)": row.get("duration_s"),
        "Chief complaint": s.get("chief_complaint"),
        "History of present illness": s.get("history_of_present_illness"),
        "Symptoms": _join(s.get("symptoms")),
        "Vitals & exam": _join(s.get("vitals_and_exam")),
        "Past history": _join(s.get("past_history")),
        "Current medications": _join(s.get("current_medications")),
        "Assessment": _join(s.get("assessment")),
        "Investigations": _join(s.get("investigations_ordered")),
        "Medications prescribed": med_str,
        "Advice & plan": _join(s.get("advice_and_plan")),
        "Follow-up": s.get("follow_up"),
        "Red flags": _join(s.get("red_flags")),
        "Patient summary": s.get("patient_summary"),
        "Transcript": transcript[:_CELL_MAX],
        "Model": row.get("model"),
    }


_FLATTEN = {"intake": _flatten_intake, "consultations": _flatten_consultation}

# Fields the free-text search scans, per dataset (already flattened keys).
_SEARCH_FIELDS = {
    "intake": ("Full name", "Phone"),
    "consultations": ("Chief complaint", "Patient summary"),
}


# --- fetch + filter ----------------------------------------------------------


def _end_of_day(value: str) -> str:
    """A bare YYYY-MM-DD `to` bound should include that whole day."""
    if len(value) == 10:
        return value + "T23:59:59.999999+00:00"
    return value


def _fetch(dataset: str, filters: ExportFilters) -> list[dict]:
    table = _TABLE[dataset]
    query = get_client().table(table).select("*").order("created_at", desc=True).limit(10000)

    if filters.date_from:
        query = query.gte("created_at", filters.date_from)
    if filters.date_to:
        query = query.lte("created_at", _end_of_day(filters.date_to))
    if filters.statuses:
        query = query.in_("status", filters.statuses)
    if filters.ids:
        query = query.in_("id", filters.ids)

    rows = query.execute().data or []

    # Filters PostgREST can't express cleanly are applied here, on already
    # date/status/id-narrowed rows.
    if dataset == "intake" and filters.flagged_only:
        rows = [r for r in rows if (r.get("flags") or [])]

    if filters.search:
        needle = filters.search.lower()
        flat_fields = _SEARCH_FIELDS[dataset]

        def matches(raw: dict) -> bool:
            flat = _FLATTEN[dataset](raw)
            return any(needle in str(flat.get(f) or "").lower() for f in flat_fields)

        rows = [r for r in rows if matches(r)]

    return rows


# --- public API --------------------------------------------------------------


def list_rows(dataset: str, filters: ExportFilters) -> list[dict]:
    """Lightweight rows for the picker + live count (not the full export)."""
    rows = _fetch(dataset, filters)
    if dataset == "intake":
        return [
            {
                "id": r.get("id"),
                "created_at": r.get("created_at"),
                "status": r.get("status"),
                "name": _get(r.get("patient") or {}, "personal.full_name"),
                "phone": _get(r.get("patient") or {}, "personal.phone"),
                "flags": len(r.get("flags") or []),
            }
            for r in rows
        ]
    return [
        {
            "id": r.get("id"),
            "created_at": r.get("created_at"),
            "status": r.get("status"),
            "title": (r.get("summary") or {}).get("chief_complaint") or "Consultation",
            "duration_s": r.get("duration_s"),
        }
        for r in rows
    ]


def _sanitize(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _to_csv(flat_rows: list[dict], headers: list[str]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in flat_rows:
        writer.writerow({h: ("" if row.get(h) is None else row.get(h)) for h in headers})
    # UTF-8 BOM so Excel opens Indian names/scripts correctly on double-click.
    return ("﻿" + buf.getvalue()).encode("utf-8")


def _to_xlsx(flat_rows: list[dict], headers: list[str], sheet: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")
    for row in flat_rows:
        ws.append([_sanitize("" if row.get(h) is None else row.get(h)) for h in headers])
    if flat_rows:
        ws.freeze_panes = "A2"
    # Width from the widest of header + a sample of the data, clamped.
    sample = flat_rows[:200]
    for i, header in enumerate(headers, 1):
        widest = max([len(str(header))] + [len(str(r.get(header) or "")) for r in sample])
        ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 2, 10), 60)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_export(dataset: str, filters: ExportFilters, fmt: str) -> tuple[bytes, str, str]:
    """Return (file_bytes, media_type, filename) for a filtered export."""
    if dataset not in DATASETS:
        raise ValueError(f"unknown dataset '{dataset}'")
    if fmt not in ("csv", "xlsx"):
        raise ValueError(f"unknown format '{fmt}'")

    rows = _fetch(dataset, filters)
    flatten = _FLATTEN[dataset]
    flat_rows = [flatten(r) for r in rows]
    # Header order comes from the flattener even when there are zero rows, so an
    # empty export is still a valid, correctly-columned file.
    headers = list(flatten({}).keys())

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"{dataset}_export_{stamp}.{fmt}"

    if fmt == "csv":
        return _to_csv(flat_rows, headers), "text/csv; charset=utf-8", filename
    sheet = "Intake forms" if dataset == "intake" else "Consultations"
    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return _to_xlsx(flat_rows, headers, sheet), media, filename
